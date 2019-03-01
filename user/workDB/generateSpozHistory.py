#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fileencoding=utf-8
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
import pymysql
import datetime
import time
pymysql.install_as_MySQLdb()

wb = Workbook()
std=wb.get_sheet_by_name('Sheet')
wb.remove_sheet(std)

alC = Alignment(horizontal="center", vertical="center")
alR = Alignment(horizontal="right", vertical="center")
alL = Alignment(horizontal="left", vertical="center")

greenFill = PatternFill(start_color='d0f5ce', end_color='d0f5ce', fill_type='solid')

ws = wb.create_sheet('Нарахування')

now = datetime.date.today()
dbAbonent = pymysql.connect(host="192.168.0.112", user="root", passwd="02071228", port=3307)
curA = dbAbonent.cursor()
date=now.strftime("%Y-%m")+'-01'
first = now.replace(day=1)
lastMonth = first - datetime.timedelta(days=1)
monthAgo = lastMonth.strftime("%Y-%m")
command='SELECT * FROM webPLA.user_private_abonent;'
curA.execute(command)
abon = curA.fetchall()
count=1
for i in abon:
    command='SELECT date, pokaz2 FROM webPLA.user_spozhistory WHERE account_id='+i[0]+';'
    curA.execute(command)
    dateHistory = curA.fetchall()
    if(str(dateHistory[len(dateHistory)-1][0])!=monthAgo+'-01'):
        pokaz1 = dateHistory[len(dateHistory)-1][1]
        command='SELECT pokazT0 FROM webPLA.user_meterdataprivate WHERE date like "'+date+'" and account_id='+i[0]+';'
        curA.execute(command)
        pokaz2 = curA.fetchall()
        print(pokaz2, i[0])
        pokaz2 = pokaz2[0][0]
        command='SELECT * FROM webPLA.user_privelege WHERE id='+str(i[9])+';'
        curA.execute(command)
        tariff = curA.fetchall()
        if(pokaz2<pokaz1):
            pokaz2=pokaz1
        different = (pokaz2-pokaz1)
        if (different<=tariff[0][2]):
            uah=different*0.9
        else:
            uah=(different-tariff[0][2])*1.68+(tariff[0][2]*0.9)
        a='A'+str(count)
        b='B'+str(count)
        c='C'+str(count)
        d='D'+str(count)
        e='E'+str(count)
        f='F'+str(count)

        cell = ws[a]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = i[0]

        cell = ws[b]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = str(monthAgo+'-01')

        cell = ws[c]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = round(pokaz1,2)

        cell = ws[d]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = round(pokaz2,2)

        cell = ws[e]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = round(different,2)

        cell = ws[f]
        cell.font = Font(size=8)
        cell.alignment = alC
        cell.value = round(uah,2)


        if(tariff[0][2]!=100):
            ws[a].fill = greenFill
            ws[b].fill = greenFill
            ws[c].fill = greenFill
            ws[d].fill = greenFill
            ws[e].fill = greenFill
            ws[f].fill = greenFill


        count=count+1
        sql = "INSERT INTO webPLA.user_spozhistory(date, pokaz1, pokaz2, different, uah, account_id) VALUES (%s, %s, %s, %s, %s, %s);"
        curA.execute(sql, (str(monthAgo+'-01'), round(pokaz1,2), round(pokaz2,2), round(different,2), round(uah,2), str(i[0])))
        dbAbonent.commit()
        sql = "INSERT INTO webPLA.user_balans(saldo, comment, date, account_id) VALUES (%s, 'Нарахування', %s, %s);"
        curA.execute(sql, (round(-uah,2), str(monthAgo+'-01'), str(i[0])))
        dbAbonent.commit()
check_name='D:\Program Files\Work\girpromenergo\Показники з будинків\\'+monthAgo+'_narah.xlsx'
wb.save(check_name)
