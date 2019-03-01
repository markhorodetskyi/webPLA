import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
import pymysql
import datetime
import time

dbAbonent = pymysql.connect(host="192.168.0.112", user="root", passwd="02071228", port=3307, charset='utf8')
curA = dbAbonent.cursor()
curA.execute("SELECT * FROM abonent.abonent_oplata;")
sal1 = curA.fetchall()


wb = Workbook()
std=wb.get_sheet_by_name('Sheet')
wb.remove_sheet(std)



for f in sal1:
    if (str(f[2])=='2018-06-01' and f[4]=='Оплата' and str(f[1]) != '110091088'):
        print(f[1])
        sql="SELECT * FROM webPLA.user_private_abonent where account_id = %s;;"
        curA.execute(sql, (f[1]))
        abon = curA.fetchall()
        for i in abon:
            ws = wb.create_sheet(i[0])
            # Printer Settings
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A5
            ws.page_margins.left = 0.30
            ws.page_margins.right = 0.30

            ws.column_dimensions['A'].width = 7.58
            ws.column_dimensions['B'].width = 8.85
            ws.column_dimensions['C'].width = 19.85
            ws.column_dimensions['D'].width = 9.01
            ws.column_dimensions['E'].width = 9.71
            ws.column_dimensions['F'].width = 13.71
            ws.column_dimensions['G'].width = 1.55181
            ws.column_dimensions['H'].width = 2.4120
            ws.column_dimensions['I'].width = 11.58
            ws.column_dimensions['J'].width = 11.71


            thin = Side(border_style="thin", color="000000")
            double = Side(border_style="double", color="000000")
            slantDashDot = Side(border_style="slantDashDot", color="000000")
            hair = Side(border_style="dotted", color="d1d1d1")

            alC = Alignment(horizontal="center", vertical="center")
            alR = Alignment(horizontal="right", vertical="center")
            alL = Alignment(horizontal="left", vertical="center")


            #---------------------------------4-------------------------------------------------
            ws.merge_cells('A4:C4')
            cell = ws['A4']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = 'Квитанція на оплату електропостачання'

            ws.merge_cells('G4:J4')
            rows = ws['G4:J4']
            for cell in rows[0]:
                cell.border = Border(top=slantDashDot, left=slantDashDot, right=slantDashDot)
            cell = ws['G4']
            cell.alignment = alC
            cell.value = 'Р/Р 26009053741970'
            #---------------------------------5-------------------------------------------------
            ws.merge_cells('G5:J5')
            rows = ws['G5:J5']
            for cell in rows[0]:
                cell.border = Border(bottom=slantDashDot, left=slantDashDot, right=slantDashDot)
            cell = ws['G5']
            cell.alignment = alC
            cell.value = 'МФО 325321, ЗКПО 41239703'
            #---------------------------------6-------------------------------------------------
            ws.merge_cells('A6:J6')
            rows = ws['A6:J6']
            for cell in rows[0]:
                cell.border = Border(bottom=double)
            cell = ws['A6']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'Одержувач платежу: ТзОВ "ДП Гірпроменерго"'

            #---------------------------------7-------------------------------------------------
            rows = ws['A7:C7']
            for cell in rows[0]:
                cell.border = Border(bottom=thin)
            cell = ws['A7']
            cell.alignment = alR
            cell.font = Font(size=8)
            cell.value = 'О/Р'

            cell = ws['B7']
            cell.font = Font(size=8)
            cell.alignment = alL

            ws.merge_cells('C7:D7')
            cell = ws['C7']
            cell.font = Font(size=8)
            cell.alignment = alL

            ws.merge_cells('E7:F7')
            cell = ws['E7']
            cell.font = Font(size=8)
            cell.alignment = alC
            #---------------------------------8-------------------------------------------------
            ws.merge_cells('A8:B8')
            rows = ws['A8:D8']
            for cell in rows[0]:
                cell.border = Border(bottom=thin)
            cell = ws['A8']
            cell.font = Font(size=8)
            cell.alignment = alL
            cell.value = 'Квитанцію виписано'

            cell = ws['C8']
            cell.font = Font(size=8)
            cell.alignment = alL

            cell = ws['E8']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Лічильник №'

            cell = ws['F8']
            cell.font = Font(size=8)
            cell.alignment = alL

            ws.merge_cells('I8:J8')
            rows = ws['I8:I15']

            cell = ws['I8']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = 'За ел.енергію, грн'

            for cell in rows[0]:
                cell.border = Border(left=thin)
            for cell in rows[1]:
                cell.border = Border(left=thin)
            for cell in rows[2]:
                cell.border = Border(left=thin)
            for cell in rows[3]:
                cell.border = Border(left=thin)
            for cell in rows[4]:
                cell.border = Border(left=thin)
            for cell in rows[5]:
                cell.border = Border(left=thin)
            for cell in rows[6]:
                cell.border = Border(left=thin)
            for cell in rows[7]:
                cell.border = Border(left=thin)
            #---------------------------------9-------------------------------------------------
            ws.merge_cells('A9:B9')
            rows = ws['A9:F9']
            for cell in rows[0]:
                cell.border = Border(bottom=thin)

            cell = ws['A9']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Спожито за'

            cell = ws['C9']
            cell.font = Font(size=8)
            cell.alignment = alL

            cell = ws['D9']
            cell.font = Font(size=8)
            cell.alignment = alR

            cell = ws['E9']
            cell.font = Font(size=8)
            cell.alignment = alL
            cell.value = 'кВт*год'

            cell = ws['J9']
            cell.font = Font(size=8)
            cell.alignment = alR
            #---------------------------------10------------------------------------------------
            ws.merge_cells('A10:A11')
            rows = ws['A10:A13']
            for cell in rows[0]:
                cell.border = Border(right=hair)
            for cell in rows[1]:
                cell.border = Border(right=hair)
            for cell in rows[2]:
                cell.border = Border(right=hair)
            for cell in rows[3]:
                cell.border = Border(right=hair)
            cell = ws['A10']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'поч. показ'

            ws.merge_cells('B10:B11')
            rows = ws['B10:B13']
            for cell in rows[0]:
                cell.border = Border(right=hair)
            for cell in rows[1]:
                cell.border = Border(right=hair)
            for cell in rows[2]:
                cell.border = Border(right=hair)
            for cell in rows[3]:
                cell.border = Border(right=hair)
            cell = ws['B10']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'кін. показ'



            ws.merge_cells('C10:C11')
            rows = ws['C10:C13']
            for cell in rows[0]:
                cell.border = Border(right=hair)
            for cell in rows[1]:
                cell.border = Border(right=hair)
            for cell in rows[2]:
                cell.border = Border(right=hair)
            for cell in rows[3]:
                cell.border = Border(right=hair)
            cell = ws['C10']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'спожито кВт*год'

            rows = ws['D12:D13']
            for cell in rows[0]:
                cell.border = Border(right=hair)
            for cell in rows[1]:
                cell.border = Border(right=hair)

            ws.merge_cells('D10:E11')
            cell = ws['D10']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'Тариф, грн'

            rows = ws['E10:E13']
            for cell in rows[0]:
                cell.border = Border(right=hair)
            for cell in rows[1]:
                cell.border = Border(right=hair)
            for cell in rows[2]:
                cell.border = Border(right=hair)
            for cell in rows[3]:
                cell.border = Border(right=hair)

            ws.merge_cells('F10:F11')
            cell = ws['F10']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alC
            cell.value = 'Сума.грн.'
            #---------------------------------11------------------------------------------------
            cell = ws['I11']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Борг:'

            cell = ws['J11']
            cell.alignment = alL
            #---------------------------------12------------------------------------------------
            cell = ws['A12']
            cell.font = Font(size=8)
            cell.alignment = alC

            cell = ws['B12']
            cell.font = Font(size=8)
            cell.alignment = alC

            cell = ws['C12']
            cell.font = Font(size=8)
            cell.alignment = alC

            cell = ws['D12']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = '(I)'

            cell = ws['E12']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = '0.9'

            cell = ws['F12']
            cell.font = Font(size=8)
            cell.alignment = alC

            cell = ws['I12']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Передплата:'

            cell = ws['J12']
            cell.alignment = alL
            #---------------------------------13------------------------------------------------
            cell = ws['D13']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = '(II)'

            cell = ws['E13']
            cell.font = Font(size=8)
            cell.alignment = alC
            cell.value = '1.68'

            cell = ws['F13']
            cell.font = Font(size=8)
            cell.alignment = alC

            cell = ws['I13']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Субсидія:'

            cell = ws['J13']
            cell.alignment = alL
            #---------------------------------14------------------------------------------------
            cell = ws['I14']
            cell.font = Font(size=8)
            cell.alignment = alR
            cell.value = 'Пільги:'

            cell = ws['J14']
            cell.alignment = alL
            #---------------------------------15------------------------------------------------
            ws.merge_cells('A15:C15')
            cell = ws['A15']
            cell.font = Font(bold=True, size=8)
            cell.alignment = alR
            cell.value = 'Тариф, грн./кВт*год:'

            ws.merge_cells('D15:G15')
            cell = ws['D15']
            cell.font = Font(size=8)
            cell.alignment = alL

            cell = ws['I15']
            cell.font = Font(bold=True)
            cell.alignment = alR
            cell.value = 'До оплати:'

            cell = ws['J15']
            cell.font = Font(bold=True)
            cell.alignment = alL

            sql="SELECT street_id FROM webPLA.user_house WHERE number = %s;"
            curA.execute(sql, (i[8]))
            house = curA.fetchall()
            for a in house:
                h_number = i[8]
                street = a[0]
                sql="SELECT town_id FROM webPLA.user_street WHERE street_name = %s;"
                curA.execute(sql, (a[0]))
                town = curA.fetchall()
                for b in town:
                    town_name = b[0]
                    adress = 'м.'+town_name+' вул.'+street+' '+h_number+' кв '+str(i[4])
                    cell = ws['E7']
                    cell.font = Font(size=8)
                    cell.value = adress
                    cell.alignment = alL
            #
            # curA.execute("SELECT * FROM webPLA.user_spozhistory WHERE account_id = %s and date = %s;")
            # history = curA.fetchall()
            osrah = str(i[0])
            cell = ws['B7']
            cell.value = osrah

            name = str(i[1])
            l_name = str(i[2])
            s_name = str(i[3])
            pib=name+' '+l_name+' '+s_name
            cell = ws['C7']
            cell.value = pib

            meter = str(i[5])
            cell = ws['F8']
            cell.value = meter

            sql="SELECT * FROM webPLA.user_privelege WHERE id = %s;"
            curA.execute(sql, (i[9]))
            limit = curA.fetchall()
            for e in limit:
                limit = e[2]
            cell = ws['D15']
            cell.value = '(I) до '+str(limit)+'кВт - 0,9грн;    (II) понад '+str(limit)+'кВт - 1,68грн.'

            now = datetime.date.today()
            first = now.replace(day=1)
            lastMonth = first - datetime.timedelta(days=1)
            today = now.strftime("%Y-%m-%d")
            monthAgo = lastMonth.strftime("%Y-%m")
            monthAgo2 = lastMonth.strftime("%m.%Y")

            cell = ws['C8']
            cell.value = today

            sql="SELECT * FROM webPLA.user_spozhistory WHERE account_id = %s and date like %s;"
            curA.execute(sql, (i[0],monthAgo+'-01'))
            history = curA.fetchall()
            for c in history:

                cell = ws['J9']
                cell.value = c[5]

                cell = ws['C9']
                cell.value = monthAgo2

                cell = ws['D9']
                cell.value = c[4]

                cell = ws['A12']
                cell.value = c[2]

                cell = ws['B12']
                cell.value = c[3]

                cell = ws['C12']
                cell.value = c[4]

                if(c[4]>100):
                    first_price = limit*0.9
                    second_price = c[5]-first_price

                    cell = ws['F12']
                    cell.value = first_price

                    cell = ws['F13']
                    cell.value = second_price
                else:
                    cell = ws['F12']
                    cell.value = c[5]

                    cell = ws['F13']
                    cell.value = 0.0

                sql="SELECT * FROM webPLA.user_balans where account_id = %s;"
                curA.execute(sql, (i[0]))
                balans = curA.fetchall()
                subsidy = 0.0
                privilege=0.0
                saldo=0
                privilege_date=monthAgo+'-01'
                for d in balans:
                    saldo = saldo + d[1]
                for d in balans:
                    if(d[2]=='Пільга' and str(d[3])==privilege_date):
                        privilege = d[1]
                        print(d[4])
                for d in balans:
                    if(d[2]=='Субсидії' and str(d[3])==privilege_date):
                        subsidy = d[1]
                        print(d[4])
                debt = saldo + c[5] - privilege - subsidy

                if(debt<0):
                    cell = ws['J11']
                    cell.value = round(debt*(-1), 2)
                    cell = ws['J12']
                    cell.value = '0'
                else:
                    cell = ws['J11']
                    cell.value = '0'
                    cell = ws['J12']
                    cell.value = round(debt, 2)

                cell = ws['J13']
                cell.value = round(subsidy, 2)
                cell = ws['J14']
                cell.value = round(privilege, 2)
                if(saldo<0):
                    cell = ws['J15']
                    cell.value = round(saldo*(-1), 2)
                else:
                    cell = ws['J15']
                    cell.value = '0.00'

                cell = ws['A16']
                cell.font = Font(size=8)
                cell.alignment = alL
                # cell.value = 'Шановний абонент! Квитанцію виписану за 2019-02-20 вважати не коректною'
                # cell = ws['A17']
                # cell.font = Font(size=8)
                # cell.alignment = alL
                # cell.value = 'Для оплати використовувати цю квитанцію (виписану за 2019-02-22). Вибачте за незручності!'
                # cell = ws['A18']
                # cell.font = Font(size=8)
                # cell.alignment = alL
                # cell.value = 'У випадку, якщо олата вже була здійснена, кошти з акумулюються на Вашому рахунку'
# check_name='D:\Program Files\Work\girpromenergo\checks\\'+str(i[0])+'_'+monthAgo+'.xlsx'
check_name='D:\Program Files\Work\girpromenergo\checks\\'+monthAgo+'_toPrint222.xlsx'
wb.save(check_name)
