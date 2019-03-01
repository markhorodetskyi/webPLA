import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook, load_workbook
import pymysql
import datetime
import time
import sys

dbAbonent = pymysql.connect(host="213.174.22.133", user="root", passwd="02071228", port=3307, charset='utf8')
curA = dbAbonent.cursor()

now = datetime.date.today()
first = now.replace(day=1)
lastMonth = first - datetime.timedelta(days=1)
monthAgo = lastMonth.strftime("%Y-%m")+'-01'

wb = load_workbook('user/workDB/saldo.xlsx')
sheet = wb['Лист1']

buff = []
opl = []

for i in range(1, sheet.max_row + 1):
    for j in range(sheet.max_column):
        if sheet[i][j].value:
            buff.append(sheet[i][j].value)
    opl.append(buff)
    buff=[]
for i in opl:
    if(len(i)==3):
        sql = "INSERT INTO webPLA.user_balans(saldo, comment, date, account_id) VALUES (%s, %s, %s, %s);"
        curA.execute(sql, (str(i[1]), str(i[2]), str(monthAgo), str(i[0])))
        dbAbonent.commit()
        print(str(i[0]), str(i[2]), str(i[1]))
    elif(len(i)==4):
        sql = "INSERT INTO webPLA.user_balans(saldo, comment, date, account_id) VALUES (%s, %s, %s, %s);"
        curA.execute(sql, (str(i[2]), str(i[3]), str(i[1]), str(i[0])))
        dbAbonent.commit()
        print(str(i[0]), str(i[1]), str(i[2]), str(i[3]))
    else:
        print(str(i[0]), 'Нічого не оплачено')



# for cell in rows[0]:
#     cell.border = Border(top=slantDashDot, left=slantDashDot, right=slantDashDot)
